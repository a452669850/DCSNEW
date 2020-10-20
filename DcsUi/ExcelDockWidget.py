from PyQt5.QtWidgets import QTabWidget, QDockWidget, QTableView, QHeaderView, QTableWidget, QMessageBox, QAbstractItemView, QMenu
from PyQt5.QtGui import QStandardItemModel, QStandardItem, QBrush, QCursor, QColor
from PyQt5.QtCore import Qt
from DcsUi.DockCLass import NewDockWidget
from DcsUi.LabelTree import LabelTree
from utils.ClientModels import Usecase, UsecaseGroup
import openpyxl
import json
import sys
import os

tittle = ['序号', '实验步骤', '测试时间', '实际结果', '预期结果', '操作位置', '是否符合预期'] # 表格表头

class TabDockWidget(NewDockWidget):
    def __init__(self, title, parent = None):
        NewDockWidget.__init__(self, title, parent=parent)
        self.ExcelTab = QTabWidget() # 初始化dock中的选项卡tab
        self.ExcelTab.setTabsClosable(True)
        self.ExcelTab.tabCloseRequested.connect(self.closeTab)
        self.parent = parent

    def addExcel(self, path):
        # 为Dock中添加规程表格 path：规程路径
        if path:
            self.TabelView = ExcelTabelView(path, self.parent)
            self.ExcelTab.addTab (self.TabelView, '规程：' + os.path.basename(path))
            self.setWidget(self.ExcelTab)

    def addUseCase(self, usecaseName):
        # 为Dock中添加用例表格 usecaseName：用例名
        self.TabelView = UsecaseTabelView(usecaseName, self.parent)
        self.ExcelTab.addTab (self.TabelView, '用例：' + usecaseName)
        self.setWidget(self.ExcelTab)

    def addUsecaseGroup(self, groupName):
        # 为Dock中添加用例组表格 groupName：用例组名
        self.TabelView = UsecaseGroupView(groupName, self.parent)
        self.ExcelTab.addTab (self.TabelView, '用例组：' + groupName)
        self.setWidget(self.ExcelTab)

    def addTree(self):
        # 为Dock添加树结构
        self.Tree = LabelTree(self.parent.projectPath)
        self.ExcelTab.addTab(self.Tree, '工程树')
        self.setWidget(self.ExcelTab)

    def closeTab(self, index):
        # Dock中选显卡关闭事件
        if not self.parent.ProcedureThread._isWork or index != self.ExcelTab.indexOf(self.parent.ProcedureThread.procedureExcel):
            self.ExcelTab.removeTab(index)
        else:
            reply = QMessageBox.question(self, '提示', '请先退出正在运行的规程！', QMessageBox.Yes)


class ExcelTabelView(QTableView):
    def __init__(self, path, parent):
        QTableView.__init__(self)
        self.parent = parent
        if os.path.exists(path):
            self.model = QStandardItemModel(0,0) # 初始化model
            self.addExcelContent(path)
            self.setModel(self.model)
            self.resizeRowsToContents()
            self.horizontalHeader().setStretchLastSection(True) # 设置表头自动填充
            self.setEditTriggers(QAbstractItemView.NoEditTriggers) # 设置表格不可修改
            self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) # 设置表格中的高度自适应
            # self.horizontalHeader().setPropertySectionResizeMode(QHeaderView.ResizeToContents);
            self.verticalHeader().setSectionResizeMode(QHeaderView.Fixed); # 设置表格中的宽度自适应
            self.setContextMenuPolicy(Qt.CustomContextMenu) # 右键菜单
            self.customContextMenuRequested.connect(self.showContextMenu) # 菜单连接信号
            self.type = 'procedure'

            
    def addExcelContent(self,path):
        # 为表格添加内容 path：规程路径
        self.setUpdatesEnabled(False) 
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        self.rowsCon = ws.rows
        for row in ws.rows:
            self.model.appendRow([QStandardItem(str(x.value)) if x.value is not None else QStandardItem(' ') for x in row])
        self.rowsLen = ws.max_column
        self.colsLen = ws.max_row
        self.rowsCon = [x for x in self.rowsCon]
        self.setUpdatesEnabled(True) 

    def changeRowColor(self, rowIndex, res):
        # 规程运行时更改行颜色 rowIndex：行号 res：是否运行成功
        if res:
            brush = QBrush(QColor(0,255,154))
        else:
            brush = QBrush(QColor(220,20,60))
        if rowIndex < 0:
            for x in range(self.rowsLen):
                self.model.setData(self.model.index(abs(rowIndex), x), brush, Qt.BackgroundRole)
        else:
            for x in range(self.rowsLen):
                self.model.setData(self.model.index(rowIndex, x), brush, Qt.BackgroundRole)
                # self.model.setData(self.model.index(rowIndex -1, x), QBrush(Qt.gray), Qt.BackgroundRole)

    def getRowContent(self, rowIndex):
        # 获取规程中执行步骤
        try:
            rowCon =  self.model.item(rowIndex, 1).text()
        except:
            try:
                if 'STEP' in self.model.item(rowIndex, 1).text():
                    return None
            except:
                return None
            return '全部'
        return rowCon

    def showContextMenu(self):  # 创建右键菜单、
        self.contextMenu = QMenu(self)
        self.actionA = self.contextMenu.addAction('从选中行开始执行')
        self.contextMenu.popup(QCursor.pos())  # 2菜单显示的位置
        self.actionA.triggered.connect(self.actionHandler)
        self.contextMenu.show()

    def actionHandler(self):
        # 从选中行执行操作
        if self.parent.ProcedureThread._isPause and self.parent.ProcedureThread._isWork:
            self.parent.procedureRunIndex = self.currentIndex().row()
            self.parent.ProcedureThread.resume()
        else:
            self.parent.procedureRunIndex = self.currentIndex().row()
            self.parent.procedureAutoRunClicked()
        self.parent.log.infoLog(f'{self.parent.procedureRunPath}规程从{self.currentIndex().row()}行开始执行')

class UsecaseTabelView(QTableView):
    def __init__(self, usecaseName, parent):
        QTableView.__init__(self)
        self.parent = parent
        self.type = 'usecase'
        self.colIndex = 0
        self.rowIndex = 2
        self.usecaseName = usecaseName
        self.usecaseOperation = json.loads(Usecase.get_by_name(usecaseName).operation)
        self.model = QStandardItemModel(0,0)
        self.addExcelContent()
        self.setModel(self.model)
        self.resizeRowsToContents()
        self.horizontalHeader().setStretchLastSection(True)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # self.horizontalHeader().setPropertySectionResizeMode(QHeaderView.ResizeToContents);
        self.verticalHeader().setSectionResizeMode(QHeaderView.Fixed);
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.setContextMenuPolicy(Qt.CustomContextMenu) # 右键菜单
        self.customContextMenuRequested.connect(self.showContextMenu) # 菜单连接信号

            
    def addExcelContent(self): 
        # 添加用例到表格中
        self.model.setItem(0, 0, QStandardItem('测试用例'))
        self.model.setItem(0, 1, QStandardItem(self.usecaseName))
        self.model.setItem(0, 2, QStandardItem('用例编号'))
        self.model.setItem(0, 3, QStandardItem(Usecase.get_by_name(self.usecaseName).number))
        self.model.appendRow([QStandardItem(str(x)) if x is not None else QStandardItem(' ') for x in tittle])
        self.setUpdatesEnabled(False)
        for step in self.usecaseOperation:
            for opr in step:
                # print(opr)
                if 'STEP' in opr:
                    self.model.setItem(self.rowIndex, self.colIndex, QStandardItem(opr))
                    self.rowIndex += 1
                else:
                    # self.insertRow(self.rowIndex)
                    for k, v in opr[1].items():
                        v = '-' if v == '' else v
                        self.model.setItem(self.rowIndex, self.colIndex, QStandardItem(v))
                        self.colIndex += 1
                    self.rowIndex += 1
                    self.colIndex = 0
        self.colsLen = self.model.rowCount()
        self.rowsLen = 7
        self.setUpdatesEnabled(True) 

    def changeRowColor(self, rowIndex, res):
        # 更改行颜色
        if res:
            brush = QBrush(QColor(0,255,154))
        else:
            brush = QBrush(QColor(220,20,60))
        try:
            self.model.item(rowIndex, 1).text()
        except:
            # for x in range(self.rowsLen):
            #     self.model.setData(self.model.index(rowIndex -1, x), QBrush(Qt.gray), Qt.BackgroundRole)
            return
        if rowIndex < 0:
            for x in range(self.rowsLen):
                self.model.setData(self.model.index(abs(rowIndex), x), brush, Qt.BackgroundRole)
        else:
            # print([x.text() for x in self.model.takeRow(rowIndex)])
            for x in range(self.rowsLen):
                self.model.setData(self.model.index(rowIndex, x), brush, Qt.BackgroundRole)
                # self.model.setData(self.model.index(rowIndex -1, x), QBrush(Qt.gray), Qt.BackgroundRole)

    def getRowContent(self, rowIndex):
        # 获取执行步骤
        try:
            rowCon =  self.model.item(rowIndex, 1).text()
        except:
            try:
                if 'STEP' in self.model.item(rowIndex, 1).text():
                    return None
            except:
                return None
            return '全部'
        return rowCon

    def showContextMenu(self):  # 创建右键菜单、
        self.contextMenu = QMenu(self)
        self.actionA = self.contextMenu.addAction('从选中行开始执行')
        self.contextMenu.popup(QCursor.pos())  # 2菜单显示的位置
        self.actionA.triggered.connect(self.actionHandler)
        self.contextMenu.show()

    def actionHandler(self):
        if self.parent.ProcedureThread._isPause and self.parent.ProcedureThread._isWork:
            self.parent.procedureRunIndex = self.currentIndex().row()
            self.parent.ProcedureThread.resume()
        else:
            self.parent.procedureRunIndex = self.currentIndex().row()
            self.parent.procedureAutoRunClicked()
        self.parent.log.infoLog(f'{self.parent.procedureRunPath}规程从{self.currentIndex().row()}行开始执行')

class UsecaseGroupView(QTableView):
    def __init__(self, groupName, parent):
        QTableView.__init__(self)
        self.parent = parent
        self.type = 'usecasegroup'
        self.groupName = groupName
        self.colIndex = 0
        self.rowIndex = 0
        self.model = QStandardItemModel(0,0)
        self.getAllUsecase()
        self.setModel(self.model)
        self.resizeRowsToContents()
        self.horizontalHeader().setStretchLastSection(True)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
        # self.horizontalHeader().setPropertySectionResizeMode(QHeaderView.ResizeToContents)
        # self.verticalHeader().setMinimumHeight(100)
        # self.verticalHeader().setMaximumHeight(100)
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)

        self.setContextMenuPolicy(Qt.CustomContextMenu) # 右键菜单
        self.customContextMenuRequested.connect(self.showContextMenu) # 菜单连接信号

    def getAllUsecase(self):
        # 获取用例组中所包含的所有用例
        for usecaseName in json.loads(UsecaseGroup.get_by_name(self.groupName).usecase):
            self.usecaseOperation = json.loads(Usecase.get_by_name(usecaseName).operation)
            self.insertUsecase(usecaseName)

    def insertUsecase(self, usecaseName):
        # 逐个插入用例
        self.setUpdatesEnabled(False)
        self.model.setItem(self.rowIndex, 0, QStandardItem('测试用例'))
        self.model.setItem(self.rowIndex, 1, QStandardItem(usecaseName))
        self.model.setItem(self.rowIndex, 2, QStandardItem('用例编号'))
        self.model.setItem(self.rowIndex, 3, QStandardItem(Usecase.get_by_name(usecaseName).number))
        self.model.appendRow([QStandardItem(str(x)) if x is not None else QStandardItem(' ') for x in tittle])
        self.rowIndex += 2
        for step in self.usecaseOperation:
            for opr in step:
                # print(opr)
                if 'STEP' in opr:
                    self.model.setItem(self.rowIndex, self.colIndex, QStandardItem(opr))
                    self.rowIndex += 1
                else:
                    for k, v in opr[1].items():
                        self.model.setItem(self.rowIndex, self.colIndex, QStandardItem(v))
                        self.colIndex += 1
                    self.rowIndex += 1
                    self.colIndex = 0
        self.rowIndex += 1
        self.colsLen = self.model.rowCount()
        self.rowsLen = 7
        self.setUpdatesEnabled(True) 

    def changeRowColor(self, rowIndex, res):
        if res:
            brush = QBrush(QColor(0,255,154))
        else:
            brush = QBrush(QColor(220,20,60))
        try:
            self.model.item(rowIndex, 1).text()
        except:
            # for x in range(self.rowsLen):
            #     self.model.setData(self.model.index(rowIndex -1, x), QBrush(Qt.gray), Qt.BackgroundRole)
            return
        if rowIndex < 0:
            for x in range(self.rowsLen):
                self.model.setData(self.model.index(abs(rowIndex), x), brush, Qt.BackgroundRole)
        else:
            # print([x.text() for x in self.model.takeRow(rowIndex)])
            for x in range(self.rowsLen):
                self.model.setData(self.model.index(rowIndex, x), brush, Qt.BackgroundRole)
                # self.model.setData(self.model.index(rowIndex -1, x), QBrush(Qt.gray), Qt.BackgroundRole)

    def getRowContent(self, rowIndex):
        try:
            rowCon =  self.model.item(rowIndex, 1).text()
        except:
            try:
                if 'STEP' in self.model.item(rowIndex, 1).text():
                    return None
            except:
                return None
            return '全部'
        return rowCon

    def showContextMenu(self):  # 创建右键菜单、
        self.contextMenu = QMenu(self)
        self.actionA = self.contextMenu.addAction('从选中行开始执行')
        self.contextMenu.popup(QCursor.pos())  # 2菜单显示的位置
        self.actionA.triggered.connect(self.actionHandler)
        self.contextMenu.show()

    def actionHandler(self):
        if self.parent.ProcedureThread._isPause and self.parent.ProcedureThread._isWork:
            self.parent.procedureRunIndex = self.currentIndex().row()
            self.parent.ProcedureThread.resume()
        else:
            self.parent.procedureRunIndex = self.currentIndex().row()
            self.parent.procedureAutoRunClicked()
        self.parent.log.infoLog(f'{self.parent.procedureRunPath}规程从{self.currentIndex().row()}行开始执行')

def clearAllItem(tabelview):
    # 清空表格中的所有内容
    for y in range(tabelview.model.rowCount()):
        for x in range(tabelview.model.columnCount()):
            tabelview.model.setData(tabelview.model.index(y, x), QBrush(), Qt.BackgroundRole)